"""
match_suppliers.py - Westcountry Group

Fuzzy-matches free-text supplier names from Prospect CRM QuoteLines (or any
other source) against the Dimensions supplier master, returning each match
with a confidence score, an ambiguity flag, and the top alternative
candidates. Output is an Excel review spreadsheet.

Usage
-----
    python match_suppliers.py \
        --suppliers  dimensions_suppliers.csv \
        --quotelines quotelines.csv \
        --out        supplier_matches.xlsx

Inputs
------
1. Dimensions supplier master (CSV or XLSX). Two columns required, however
   named in the export -- they are auto-detected by header keyword:
       * supplier code  (e.g. "ARRW001")  -- header contains "code" or "ref"
       * supplier name  (e.g. "Arrow Group Global Ltd.") -- header contains "name"

2. QuoteLines (CSV or XLSX). At minimum a column whose header contains
   "supplier" (the free-text value). Any other columns (QuoteId, LineId,
   ProductCode, Description) are passed through to the output so the review
   spreadsheet stays useful.

Output
------
An Excel workbook with one sheet "Matches":
    | Original columns... | Matched Code | Matched Name | Confidence
    | Status | Ambiguous? | Alt 1 (code/name/score) | Alt 2 (code/name/score) | Action |

Confidence bands (tuneable via --threshold and --ambiguity-gap):
    >= 90 and gap >= 10   ->  CONFIDENT     (Action: Apply)
    >= 90 and gap <  10   ->  AMBIGUOUS     (Action: Review)
    70..89                ->  REVIEW        (Action: Review)
    <  70                 ->  NO MATCH      (Action: New supplier?)
    exact code hit        ->  EXACT (100)   (Action: Apply)
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from rapidfuzz import fuzz, process
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Normalisation
# ---------------------------------------------------------------------------

# Tokens stripped during normalisation. Kept conservative -- only legal /
# incorporation suffixes that carry no distinguishing meaning. Words like
# "Group", "Global", "International", "Holdings", "UK" are deliberately NOT
# stripped: "Arrow Group" vs "Arrow Furniture" must stay distinguishable.
_SUFFIX_TOKENS = [
    "limited", "ltd",
    "plc", "llp", "llc",
    "inc", "incorporated",
    "corp", "corporation",
    "gmbh", "ag", "sa", "srl", "bv",
    "pty",
    "company",
]

_PUNCT_RE = re.compile(r"[^\w\s]")
_WHITESPACE_RE = re.compile(r"\s+")


def normalise(name) -> str:
    """Lowercase, strip punctuation, strip legal suffixes, collapse whitespace."""
    if name is None:
        return ""
    s = str(name).lower().strip()
    s = _PUNCT_RE.sub(" ", s)
    s = _WHITESPACE_RE.sub(" ", s).strip()
    if s.startswith("the "):
        s = s[4:]
    changed = True
    while changed:
        changed = False
        for tok in _SUFFIX_TOKENS:
            pattern = rf"(?:^|\s){re.escape(tok)}(?:\s|$)"
            new = re.sub(pattern, " ", s)
            if new != s:
                s = new
                changed = True
    s = _WHITESPACE_RE.sub(" ", s).strip()
    return s


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def find_column(df: pd.DataFrame, keywords, required: bool = True):
    """Find a column whose name (lowercased) contains any of the keywords."""
    lc = {c.lower(): c for c in df.columns}
    for kw in keywords:
        for lc_name, orig in lc.items():
            if kw in lc_name:
                return orig
    if required:
        raise ValueError(
            f"Could not find a column matching any of {keywords} in "
            f"columns: {list(df.columns)}"
        )
    return None


def read_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return pd.read_excel(path)
    if suffix in {".csv", ".tsv", ".txt"}:
        sep = "\t" if suffix == ".tsv" else ","
        return pd.read_csv(path, sep=sep, dtype=str, keep_default_na=False)
    raise ValueError(f"Unsupported file type: {path.suffix}")


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

@dataclass
class Supplier:
    code: str
    name: str
    norm: str


@dataclass
class MatchResult:
    original: str
    code: str
    name: str
    score: float
    status: str               # EXACT | CONFIDENT | AMBIGUOUS | REVIEW | NO MATCH | MISSING
    ambiguous: bool
    alternatives: list        # list of (code, name, score)
    action: str


def build_supplier_index(df, code_col, name_col):
    suppliers = []
    seen_codes = set()
    for _, row in df.iterrows():
        code = str(row[code_col]).strip()
        name = str(row[name_col]).strip()
        if not code or not name or code.lower() == "nan" or name.lower() == "nan":
            continue
        if code in seen_codes:
            continue
        seen_codes.add(code)
        suppliers.append(Supplier(code=code, name=name, norm=normalise(name)))
    return suppliers


def match_one(text, suppliers, code_lookup, threshold, ambiguity_gap):
    if text is None or not str(text).strip():
        return MatchResult(
            original=str(text) if text is not None else "",
            code="", name="", score=0.0,
            status="MISSING", ambiguous=False, alternatives=[],
            action="Add supplier text",
        )

    raw = str(text).strip()

    # 1. Exact code already entered?
    if raw.upper() in code_lookup:
        s = code_lookup[raw.upper()]
        return MatchResult(
            original=raw, code=s.code, name=s.name, score=100.0,
            status="EXACT", ambiguous=False, alternatives=[], action="Apply",
        )

    norm = normalise(raw)
    if not norm:
        return MatchResult(
            original=raw, code="", name="", score=0.0,
            status="NO MATCH", ambiguous=False, alternatives=[],
            action="New supplier?",
        )

    # 2. Fuzzy match using WRatio against normalised supplier names.
    choices = {i: s.norm for i, s in enumerate(suppliers)}
    extracted = process.extract(norm, choices, scorer=fuzz.WRatio, limit=5)
    if not extracted:
        return MatchResult(
            original=raw, code="", name="", score=0.0,
            status="NO MATCH", ambiguous=False, alternatives=[],
            action="New supplier?",
        )

    rescored = []
    for _, score, idx in extracted:
        s = suppliers[idx]
        # Promote to a perfect score when normalised forms are identical
        # (WRatio occasionally lands at 98-99 on identical strings).
        if s.norm == norm:
            score = 100.0
        rescored.append((score, s))
    rescored.sort(key=lambda t: t[0], reverse=True)

    best_score, best = rescored[0]
    best_score = min(best_score, 100.0)
    second_score = rescored[1][0] if len(rescored) > 1 else 0.0

    alts = [(r[1].code, r[1].name, round(min(r[0], 100.0), 1))
            for r in rescored[1:3]]

    if best_score >= threshold and (best_score - second_score) >= ambiguity_gap:
        return MatchResult(
            original=raw, code=best.code, name=best.name,
            score=round(best_score, 1),
            status="CONFIDENT", ambiguous=False, alternatives=alts, action="Apply",
        )
    if best_score >= threshold:
        return MatchResult(
            original=raw, code=best.code, name=best.name,
            score=round(best_score, 1),
            status="AMBIGUOUS", ambiguous=True, alternatives=alts, action="Review",
        )
    if best_score >= 70:
        return MatchResult(
            original=raw, code=best.code, name=best.name,
            score=round(best_score, 1),
            status="REVIEW", ambiguous=False, alternatives=alts, action="Review",
        )
    return MatchResult(
        original=raw, code="", name="", score=round(best_score, 1),
        status="NO MATCH", ambiguous=False, alternatives=alts,
        action="New supplier?",
    )


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

STATUS_FILL = {
    "EXACT":     PatternFill("solid", fgColor="C6EFCE"),  # green
    "CONFIDENT": PatternFill("solid", fgColor="DDEBF7"),  # blue
    "AMBIGUOUS": PatternFill("solid", fgColor="FFEB9C"),  # amber
    "REVIEW":    PatternFill("solid", fgColor="FFE699"),  # light amber
    "NO MATCH":  PatternFill("solid", fgColor="FFC7CE"),  # red
    "MISSING":   PatternFill("solid", fgColor="D9D9D9"),  # grey
}


def write_excel(df_in, results, supplier_text_col, out_path, mfr_ref_col=None):
    out_rows = []
    for orig_row, r in zip(df_in.to_dict(orient="records"), results):
        row = dict(orig_row)
        row["Dims Code"] = r.code
        row["Dims Supplier Name"] = r.name
        row["Confidence"] = r.score
        row["Status"] = r.status
        row["Ambiguous?"] = "Y" if r.ambiguous else ""
        for i in range(2):
            if i < len(r.alternatives):
                ac, an, asc = r.alternatives[i]
                row[f"Alt {i+1} Code"] = ac
                row[f"Alt {i+1} Name"] = an
                row[f"Alt {i+1} Score"] = asc
            else:
                row[f"Alt {i+1} Code"] = ""
                row[f"Alt {i+1} Name"] = ""
                row[f"Alt {i+1} Score"] = ""
        row["Action"] = r.action
        out_rows.append(row)

    out_df = pd.DataFrame(out_rows)

    # Move free-text supplier column next to Dims Code for easy review
    cols = list(out_df.columns)
    if supplier_text_col in cols and "Dims Code" in cols:
        cols.remove(supplier_text_col)
        insert_at = cols.index("Dims Code")
        cols.insert(insert_at, supplier_text_col)
        out_df = out_df[cols]

    # If a manufacturer-reference column was named, rename it to
    # "Supplier Code (Mfr Ref)" and position it directly after Dims Code
    # so the reviewer sees both identifiers side-by-side. They are
    # SUPPOSED to differ (different identifier systems) — this is not a
    # mismatch flag, just the full picture for the human reviewer.
    if mfr_ref_col and mfr_ref_col in out_df.columns:
        out_df = out_df.rename(columns={mfr_ref_col: "Supplier Code (Mfr Ref)"})
        cols = list(out_df.columns)
        cols.remove("Supplier Code (Mfr Ref)")
        insert_at = cols.index("Dims Code") + 1
        cols.insert(insert_at, "Supplier Code (Mfr Ref)")
        out_df = out_df[cols]

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Matches", index=False)
        ws = writer.sheets["Matches"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.freeze_panes = "A2"

        status_col_idx = list(out_df.columns).index("Status") + 1
        for row_i in range(2, len(out_df) + 2):
            status = ws.cell(row=row_i, column=status_col_idx).value
            fill = STATUS_FILL.get(status)
            if fill is None:
                continue
            for col_i in range(1, len(out_df.columns) + 1):
                ws.cell(row=row_i, column=col_i).fill = fill

        for col_i, col in enumerate(out_df.columns, start=1):
            values = [str(v) for v in out_df[col].astype(str).tolist()] + [str(col)]
            width = min(max(len(v) for v in values) + 2, 60)
            ws.column_dimensions[get_column_letter(col_i)].width = width

        summary = out_df.groupby("Status").size().rename("Count").reset_index()
        summary.to_excel(writer, sheet_name="Summary", index=False)
        s = writer.sheets["Summary"]
        for cell in s[1]:
            cell.font = header_font
            cell.fill = header_fill
        s.column_dimensions["A"].width = 18
        s.column_dimensions["B"].width = 10


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main(argv=None):
    p = argparse.ArgumentParser(description="Match free-text supplier names against Dimensions supplier master")
    p.add_argument("--suppliers", required=True, type=Path,
                   help="Dimensions supplier master CSV/XLSX")
    p.add_argument("--quotelines", required=True, type=Path,
                   help="QuoteLines CSV/XLSX containing the free-text supplier column")
    p.add_argument("--out", required=True, type=Path, help="Output XLSX path")
    p.add_argument("--supplier-col", default=None,
                   help="Override column name for the free-text supplier in quotelines")
    p.add_argument("--mfr-ref-col", default=None,
                   help="Column name in quotelines holding the manufacturer's "
                        "part reference (the 'Supplier Code' field on ProductItem "
                        "as WCG uses it). When set, the column is renamed to "
                        "'Supplier Code (Mfr Ref)' and positioned next to Dims "
                        "Code in the output for side-by-side review.")
    p.add_argument("--threshold", type=int, default=90,
                   help="Confidence threshold (default 90)")
    p.add_argument("--ambiguity-gap", type=int, default=10,
                   help="Gap between best and second-best to count as unambiguous (default 10)")
    args = p.parse_args(argv)

    sup_df = read_table(args.suppliers)
    qline_df = read_table(args.quotelines)

    sup_code_col = find_column(sup_df, ["code", "ref"])
    sup_name_col = find_column(sup_df, ["name", "supplier"])
    if sup_code_col == sup_name_col:
        non_code = [c for c in sup_df.columns if c != sup_code_col]
        if non_code:
            sup_name_col = non_code[0]

    supplier_text_col = args.supplier_col or find_column(qline_df, ["supplier"], required=True)

    suppliers = build_supplier_index(sup_df, sup_code_col, sup_name_col)
    code_lookup = {s.code.upper(): s for s in suppliers}

    print(f"Loaded {len(suppliers)} suppliers from {args.suppliers.name}")
    print(f"Loaded {len(qline_df)} quote lines from {args.quotelines.name}")
    print(f"Matching against free-text column: {supplier_text_col!r}")

    results = [
        match_one(t, suppliers, code_lookup,
                  threshold=args.threshold,
                  ambiguity_gap=args.ambiguity_gap)
        for t in qline_df[supplier_text_col].tolist()
    ]

    args.out.parent.mkdir(parents=True, exist_ok=True)
    write_excel(qline_df, results, supplier_text_col, args.out,
                mfr_ref_col=args.mfr_ref_col)

    counts = Counter(r.status for r in results)
    print()
    print("Match summary:")
    for status in ["EXACT", "CONFIDENT", "AMBIGUOUS", "REVIEW", "NO MATCH", "MISSING"]:
        print(f"  {status:<10} {counts.get(status, 0)}")
    print()
    print(f"Wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
